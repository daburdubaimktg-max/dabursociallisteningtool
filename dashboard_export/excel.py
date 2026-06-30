"""Build the live competitor-growth Excel dashboard from the follower store.

Sheets:
  Overview     headline KPIs + top movers (IG & TikTok)
  Instagram    brand × month follower matrix (+ latest MoM delta / %)
  TikTok       brand × month follower matrix (+ latest MoM delta / %)
  Growth       MoM-growth leaderboard, both platforms
  Combined     IG vs TikTok latest followers side by side (separate columns — Rule 2)
  By Category  total followers per category, per platform
  By Region    total followers per region, per platform
  Trends       native line chart of the top brands' follower trajectories

Reach only: no sentiment is read or written here.
"""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from growth.models import GrowthPlatform
from growth.store import GrowthStore
from kpi.growth import combined_latest, leaderboard

# ---- palette ---------------------------------------------------------------------
_NAVY = "1F3864"
_BLUE = "2E5496"
_LIGHT = "D9E1F2"
_GREEN = "C6EFCE"
_RED = "FFC7CE"
_GREY = "808080"

_HDR_FILL = PatternFill("solid", fgColor=_NAVY)
_SUB_FILL = PatternFill("solid", fgColor=_BLUE)
_BAND_FILL = PatternFill("solid", fgColor=_LIGHT)
_POS_FILL = PatternFill("solid", fgColor=_GREEN)
_NEG_FILL = PatternFill("solid", fgColor=_RED)
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=16, color=_NAVY)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_INT_FMT = "#,##0"
_PCT_FMT = "0.0%"


def _style_header(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _autosize(ws: Worksheet, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ---- per-platform matrix ---------------------------------------------------------
def _matrix_sheet(ws: Worksheet, store: GrowthStore, platform: GrowthPlatform) -> None:
    periods = store.periods(platform)
    series = store.series_by_brand(platform)
    meta = {s.brand: (s.category, s.region) for s in store.snapshots(platform)}

    ws.cell(1, 1, f"{platform.value.title()} — Competitor Follower Growth").font = _TITLE_FONT
    header = ["Brand", "Category", "Region", *periods, "Latest", "MoM Δ", "MoM %"]
    hrow = 3
    for c, h in enumerate(header, start=1):
        ws.cell(hrow, c, h)
    _style_header(ws, hrow, len(header))
    ws.freeze_panes = ws.cell(hrow + 1, 4)

    # brands sorted by latest follower count desc
    def latest(b: str) -> int:
        s = series[b]
        return s[max(s)] if s else 0

    brands = sorted(series, key=latest, reverse=True)
    r = hrow + 1
    for i, brand in enumerate(brands):
        cat, region = meta.get(brand, (None, None))
        ws.cell(r, 1, brand)
        ws.cell(r, 2, cat)
        ws.cell(r, 3, region)
        s = series[brand]
        for c, period in enumerate(periods, start=4):
            v = s.get(period)
            cell = ws.cell(r, c, v)
            cell.number_format = _INT_FMT
        ordered = sorted(s)
        last_v = s[ordered[-1]] if ordered else None
        prev_v = s[ordered[-2]] if len(ordered) >= 2 else None
        lc = 4 + len(periods)
        ws.cell(r, lc, last_v).number_format = _INT_FMT
        if last_v is not None and prev_v is not None:
            delta = last_v - prev_v
            fill = _POS_FILL if delta > 0 else (_NEG_FILL if delta < 0 else _BAND_FILL)
            dcell = ws.cell(r, lc + 1, delta)
            dcell.number_format = _INT_FMT
            dcell.fill = fill
            pcell = ws.cell(r, lc + 2, round(delta / prev_v, 4) if prev_v else None)
            pcell.number_format = _PCT_FMT
            pcell.fill = fill
        if i % 2 == 1:
            for c in range(1, 4):
                ws.cell(r, c).fill = _BAND_FILL
        r += 1

    _autosize(ws, {1: 26, 2: 16, 3: 14, **{c: 11 for c in range(4, len(header) + 1)}})


# ---- growth leaderboard ----------------------------------------------------------
def _growth_sheet(ws: Worksheet, store: GrowthStore) -> None:
    ws.cell(1, 1, "Follower Growth Leaderboard (window: first → latest period)").font = _TITLE_FONT
    header = [
        "Platform",
        "Brand",
        "Category",
        "Region",
        "From",
        "To",
        "Start",
        "End",
        "Gain (Δ)",
        "Growth %",
    ]
    hrow = 3
    for c, h in enumerate(header, start=1):
        ws.cell(hrow, c, h)
    _style_header(ws, hrow, len(header))
    ws.freeze_panes = ws.cell(hrow + 1, 1)

    r = hrow + 1
    for platform in (GrowthPlatform.instagram, GrowthPlatform.tiktok):
        for row in leaderboard(store, platform):
            ws.cell(r, 1, platform.value)
            ws.cell(r, 2, row.brand)
            ws.cell(r, 3, row.category)
            ws.cell(r, 4, row.region)
            ws.cell(r, 5, row.start_period)
            ws.cell(r, 6, row.end_period)
            ws.cell(r, 7, row.start_followers).number_format = _INT_FMT
            ws.cell(r, 8, row.end_followers).number_format = _INT_FMT
            fill = _POS_FILL if row.delta > 0 else (_NEG_FILL if row.delta < 0 else _BAND_FILL)
            dcell = ws.cell(r, 9, row.delta)
            dcell.number_format = _INT_FMT
            dcell.fill = fill
            pcell = ws.cell(r, 10, row.pct_change)
            pcell.number_format = _PCT_FMT
            pcell.fill = fill
            r += 1
    _autosize(ws, {1: 11, 2: 26, 3: 16, 4: 14, 5: 9, 6: 9, 7: 11, 8: 11, 9: 12, 10: 11})


# ---- combined IG vs TikTok -------------------------------------------------------
def _combined_sheet(ws: Worksheet, store: GrowthStore) -> None:
    ws.cell(1, 1, "Combined — Instagram vs TikTok (latest followers)").font = _TITLE_FONT
    ws.cell(
        2, 1, "IG and TikTok shown as separate columns — never blended into one score."
    ).font = Font(italic=True, color=_GREY)
    header = ["Brand", "Instagram", "TikTok", "Total Audience"]
    hrow = 4
    for c, h in enumerate(header, start=1):
        ws.cell(hrow, c, h)
    _style_header(ws, hrow, len(header))
    ws.freeze_panes = ws.cell(hrow + 1, 1)

    r = hrow + 1
    for i, row in enumerate(combined_latest(store)):
        ws.cell(r, 1, row["brand"])
        ws.cell(r, 2, row["instagram"]).number_format = _INT_FMT
        ws.cell(r, 3, row["tiktok"]).number_format = _INT_FMT
        ws.cell(r, 4, row["total"]).number_format = _INT_FMT
        if i % 2 == 1:
            for c in range(1, 5):
                ws.cell(r, c).fill = _BAND_FILL
        r += 1
    _autosize(ws, {1: 26, 2: 14, 3: 14, 4: 16})


# ---- category / region rollups ---------------------------------------------------
def _rollup_sheet(ws: Worksheet, store: GrowthStore, dim: str, title: str) -> None:
    ws.cell(1, 1, title).font = _TITLE_FONT
    header = [dim.title(), "Instagram", "TikTok", "Total", "Brands"]
    hrow = 3
    for c, h in enumerate(header, start=1):
        ws.cell(hrow, c, h)
    _style_header(ws, hrow, len(header))

    def latest_by_dim(platform: GrowthPlatform) -> tuple[dict[str, int], dict[str, set]]:
        totals: dict[str, int] = {}
        brands: dict[str, set] = {}
        snaps = store.snapshots(platform)
        # latest period per (brand) then group by dim value
        series = store.series_by_brand(platform)
        meta = {s.brand: getattr(s, dim) for s in snaps}
        for brand, s in series.items():
            if not s:
                continue
            val = meta.get(brand) or "Unspecified"
            totals[val] = totals.get(val, 0) + s[max(s)]
            brands.setdefault(val, set()).add(brand)
        return totals, brands

    ig_t, ig_b = latest_by_dim(GrowthPlatform.instagram)
    tt_t, tt_b = latest_by_dim(GrowthPlatform.tiktok)
    keys = sorted(
        set(ig_t) | set(tt_t), key=lambda k: ig_t.get(k, 0) + tt_t.get(k, 0), reverse=True
    )
    r = hrow + 1
    for i, k in enumerate(keys):
        ws.cell(r, 1, k)
        ws.cell(r, 2, ig_t.get(k, 0)).number_format = _INT_FMT
        ws.cell(r, 3, tt_t.get(k, 0)).number_format = _INT_FMT
        ws.cell(r, 4, ig_t.get(k, 0) + tt_t.get(k, 0)).number_format = _INT_FMT
        ws.cell(r, 5, len(ig_b.get(k, set()) | tt_b.get(k, set())))
        if i % 2 == 1:
            for c in range(1, 6):
                ws.cell(r, c).fill = _BAND_FILL
        r += 1
    _autosize(ws, {1: 22, 2: 14, 3: 14, 4: 14, 5: 8})


# ---- overview --------------------------------------------------------------------
def _overview_sheet(ws: Worksheet, store: GrowthStore) -> None:
    ws.cell(1, 1, "Competitor Social Media Growth — Live Dashboard").font = Font(
        bold=True, size=18, color=_NAVY
    )
    ws.cell(
        2, 1, "Reach dimension only (follower growth). Sentiment is tracked separately."
    ).font = Font(italic=True, color=_GREY)

    ig_periods = store.periods(GrowthPlatform.instagram)
    tt_periods = store.periods(GrowthPlatform.tiktok)
    ig_snaps = store.snapshots(GrowthPlatform.instagram)
    tt_snaps = store.snapshots(GrowthPlatform.tiktok)

    kpis = [
        ("Instagram brands tracked", len({s.brand for s in ig_snaps})),
        ("TikTok brands tracked", len({s.brand for s in tt_snaps})),
        ("Instagram periods", f"{ig_periods[0]} → {ig_periods[-1]}" if ig_periods else "—"),
        ("TikTok periods", f"{tt_periods[0]} → {tt_periods[-1]}" if tt_periods else "—"),
        ("Total follower snapshots", store.snapshot_count()),
    ]
    r = 4
    for label, val in kpis:
        lc = ws.cell(r, 1, label)
        lc.font = Font(bold=True, color=_NAVY)
        lc.fill = _BAND_FILL
        ws.cell(r, 2, val)
        r += 1

    # top movers
    r += 1
    ws.cell(r, 1, "Top 5 movers (absolute follower gain)").font = Font(
        bold=True, size=13, color=_NAVY
    )
    r += 1
    head = ["Platform", "Brand", "Gain (Δ)", "Growth %"]
    for c, h in enumerate(head, start=1):
        ws.cell(r, c, h)
    _style_header(ws, r, len(head))
    r += 1
    for platform in (GrowthPlatform.instagram, GrowthPlatform.tiktok):
        for row in leaderboard(store, platform)[:5]:
            ws.cell(r, 1, platform.value)
            ws.cell(r, 2, row.brand)
            ws.cell(r, 3, row.delta).number_format = _INT_FMT
            ws.cell(r, 4, row.pct_change).number_format = _PCT_FMT
            r += 1
    _autosize(ws, {1: 26, 2: 26, 3: 14, 4: 12})


# ---- trends chart ----------------------------------------------------------------
def _trends_sheet(
    ws: Worksheet, store: GrowthStore, platform: GrowthPlatform, top_n: int = 8
) -> None:
    periods = store.periods(platform)
    series = store.series_by_brand(platform)
    if not periods or not series:
        ws.cell(1, 1, f"No {platform.value} data").font = _TITLE_FONT
        return

    def latest(b: str) -> int:
        s = series[b]
        return s[max(s)] if s else 0

    brands = sorted(series, key=latest, reverse=True)[:top_n]
    ws.cell(
        1, 1, f"{platform.value.title()} follower trajectories — top {len(brands)} brands"
    ).font = _TITLE_FONT

    # data table feeding the chart (periods down column A, brands across)
    hrow = 3
    ws.cell(hrow, 1, "Period")
    for c, brand in enumerate(brands, start=2):
        ws.cell(hrow, c, brand)
    _style_header(ws, hrow, 1 + len(brands))
    for ri, period in enumerate(periods, start=hrow + 1):
        ws.cell(ri, 1, period)
        for c, brand in enumerate(brands, start=2):
            v = series[brand].get(period)
            ws.cell(ri, c, v).number_format = _INT_FMT

    chart = LineChart()
    chart.title = f"{platform.value.title()} — Follower Growth"
    chart.style = 12
    chart.y_axis.title = "Followers"
    chart.x_axis.title = "Month"
    chart.height = 11
    chart.width = 26
    data = Reference(
        ws, min_col=2, max_col=1 + len(brands), min_row=hrow, max_row=hrow + len(periods)
    )
    cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=hrow + len(periods))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{get_column_letter(3 + len(brands))}3")
    _autosize(ws, {1: 10, **{c: 12 for c in range(2, 2 + len(brands))}})


# ---- entrypoint ------------------------------------------------------------------
def build_workbook(store: GrowthStore) -> Workbook:
    wb = Workbook()
    _overview_sheet(wb.active, store)
    wb.active.title = "Overview"
    _matrix_sheet(wb.create_sheet("Instagram"), store, GrowthPlatform.instagram)
    _matrix_sheet(wb.create_sheet("TikTok"), store, GrowthPlatform.tiktok)
    _growth_sheet(wb.create_sheet("Growth"), store)
    _combined_sheet(wb.create_sheet("Combined"), store)
    _rollup_sheet(wb.create_sheet("By Category"), store, "category", "Followers by Category")
    _rollup_sheet(wb.create_sheet("By Region"), store, "region", "Followers by Region")
    _trends_sheet(wb.create_sheet("IG Trends"), store, GrowthPlatform.instagram)
    _trends_sheet(wb.create_sheet("TikTok Trends"), store, GrowthPlatform.tiktok)
    return wb


def workbook_bytes(store: GrowthStore) -> bytes:
    buf = io.BytesIO()
    build_workbook(store).save(buf)
    return buf.getvalue()


def write_dashboard(store: GrowthStore, path: str | Path) -> Path:
    path = Path(path)
    build_workbook(store).save(path)
    return path
