"""End-to-end: seed → collect → build the Excel dashboard.

Mirrors the slice philosophy — one full vertical path, exercised over the bundled
seed data, with the workbook asserted to be well-formed.
"""

from __future__ import annotations

import io

import openpyxl

from adapters.profile.adapter import FollowerAdapter
from adapters.profile.provider import SeedFollowerProvider
from dashboard_export.excel import build_workbook, workbook_bytes
from growth.config import load_competitors
from growth.models import GrowthPlatform
from growth.pipeline import collect_period, seed_history
from growth.store import GrowthStore


def test_seed_populates_both_platforms():
    store = GrowthStore()
    n = seed_history(store)
    assert n > 1000  # the bundled tracker has a couple thousand snapshots
    assert store.periods(GrowthPlatform.instagram)
    assert store.periods(GrowthPlatform.tiktok)


def test_collect_period_is_idempotent():
    store = GrowthStore()
    seed_history(store)
    before = store.snapshot_count()
    # collecting an already-present period must not create duplicates
    period = store.periods(GrowthPlatform.instagram)[-1]
    collect_period(store, period)
    collect_period(store, period)
    assert store.snapshot_count() == before


def test_adapter_skips_brands_without_a_handle_or_reading():
    comps = load_competitors()
    adapter = FollowerAdapter(provider=SeedFollowerProvider(period="2026-06"))
    snaps = adapter.collect(comps, GrowthPlatform.tiktok, "2026-06")
    # every emitted snapshot has a handle and the requested period
    assert snaps
    assert all(s.period == "2026-06" and s.handle for s in snaps)


def test_build_workbook_has_expected_sheets_and_chart():
    store = GrowthStore()
    seed_history(store)
    wb = build_workbook(store)
    expected = {
        "Overview",
        "Instagram",
        "TikTok",
        "Growth",
        "Combined",
        "By Category",
        "By Region",
        "IG Trends",
        "TikTok Trends",
    }
    assert expected.issubset(set(wb.sheetnames))
    # the trends sheets carry native charts
    assert wb["IG Trends"]._charts
    assert wb["TikTok Trends"]._charts


def test_workbook_bytes_reopen_and_have_data_rows():
    store = GrowthStore()
    seed_history(store)
    data = workbook_bytes(store)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ig = wb["Instagram"]
    # header on row 3, at least a handful of brand rows below it
    assert ig.cell(3, 1).value == "Brand"
    assert ig.max_row > 10
